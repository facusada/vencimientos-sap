from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from app.services.workbook_merge_service import merge_ewa_workbooks


def test_merge_ewa_workbooks_concatenates_base_rows_in_input_order():
    first_workbook = _build_workbook(
        [
            ["Cliente", "Componente", "FechaVencimiento"],
            ["ACES", "SQL SERVER 2012", "2022-07-12"],
            ["ACES", "Windows Server 2012 R2", "2023-10-10"],
        ]
    )
    second_workbook = _build_workbook(
        [
            ["Cliente", "Componente", "FechaVencimiento"],
            ["MRP", "SAP S/4HANA PRIVATE CLOUD 2023", "2030-12-31"],
            ["MRP", "SAP ACCESS CONTROL 12.0", "2027-12-31"],
        ]
    )

    payload = merge_ewa_workbooks(
        [
            ("ACES_vencimientos_Mayo_2026.xlsx", first_workbook),
            ("MRP_vencimientos_Mayo_2026.xlsx", second_workbook),
        ]
    )

    workbook = load_workbook(filename=BytesIO(payload))
    assert workbook.sheetnames == ["Base"]

    sheet = workbook["Base"]
    assert [sheet.cell(row=1, column=column).value for column in range(1, 4)] == [
        "Cliente",
        "Componente",
        "FechaVencimiento",
    ]
    assert [sheet.cell(row=2, column=column).value for column in range(1, 4)] == [
        "ACES",
        "SQL SERVER 2012",
        "2022-07-12",
    ]
    assert [sheet.cell(row=3, column=column).value for column in range(1, 4)] == [
        "ACES",
        "Windows Server 2012 R2",
        "2023-10-10",
    ]
    assert [sheet.cell(row=4, column=column).value for column in range(1, 4)] == [
        "MRP",
        "SAP S/4HANA PRIVATE CLOUD 2023",
        "2030-12-31",
    ]
    assert [sheet.cell(row=5, column=column).value for column in range(1, 4)] == [
        "MRP",
        "SAP ACCESS CONTROL 12.0",
        "2027-12-31",
    ]


def test_merge_ewa_workbooks_requires_at_least_two_workbooks():
    workbook = _build_workbook(
        [
            ["Cliente", "Componente", "FechaVencimiento"],
            ["ACES", "SQL SERVER 2012", "2022-07-12"],
        ]
    )

    with pytest.raises(ValueError, match="al menos 2 workbooks"):
        merge_ewa_workbooks([("one.xlsx", workbook)])


def test_merge_ewa_workbooks_rejects_missing_base_sheet():
    workbook = Workbook()
    workbook.active.title = "OtraHoja"
    workbook.active.append(["Cliente", "Componente", "FechaVencimiento"])

    buffer = BytesIO()
    workbook.save(buffer)

    valid_workbook = _build_workbook(
        [
            ["Cliente", "Componente", "FechaVencimiento"],
            ["MRP", "SAP ACCESS CONTROL 12.0", "2027-12-31"],
        ]
    )

    with pytest.raises(ValueError, match="hoja Base"):
        merge_ewa_workbooks(
            [
                ("invalid.xlsx", buffer.getvalue()),
                ("valid.xlsx", valid_workbook),
            ]
        )


def test_merge_ewa_workbooks_rejects_invalid_base_header():
    invalid_workbook = _build_workbook(
        [
            ["Cliente", "Nombre", "Fecha"],
            ["ACES", "SQL SERVER 2012", "2022-07-12"],
        ]
    )
    valid_workbook = _build_workbook(
        [
            ["Cliente", "Componente", "FechaVencimiento"],
            ["MRP", "SAP ACCESS CONTROL 12.0", "2027-12-31"],
        ]
    )

    with pytest.raises(ValueError, match="columnas requeridas"):
        merge_ewa_workbooks(
            [
                ("invalid.xlsx", invalid_workbook),
                ("valid.xlsx", valid_workbook),
            ]
        )


def _build_workbook(rows: list[list[str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Base"

    for row in rows:
        sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
