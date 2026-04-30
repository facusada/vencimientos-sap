from io import BytesIO
from datetime import date

from openpyxl import load_workbook

from app.models.expiration import ExpirationRecord
from app.models.expiration import ConsolidatedExpiration
from app.models.expiration import ConsolidatedWorkbookData
from app.services.excel_service import build_expiration_workbook
from app.services.excel_service import build_consolidated_workbook


def test_build_expiration_workbook_generates_expected_columns_and_rows():
    records = [
        ExpirationRecord(
            source_section="SAP Kernel Release",
            name="Kernel",
            expiration_date="2026-12-31",
            milestone="",
        ),
        ExpirationRecord(
            source_section="Certificates",
            name="Certificate ABC",
            expiration_date="2026-10-01",
            milestone="End of Extended Vendor Support",
        ),
    ]

    payload = build_expiration_workbook(records)

    workbook = load_workbook(filename=BytesIO(payload))
    sheet = workbook.active

    assert sheet.title == "Expirations"
    assert sheet["A1"].value == "Seccion"
    assert sheet["B1"].value == "Nombre"
    assert sheet["C1"].value == "Hito"
    assert sheet["D1"].value == "Fecha"
    assert sheet["A2"].value == "SAP Kernel Release"
    assert sheet["B2"].value == "Kernel"
    assert sheet["C2"].value is None
    assert sheet["D2"].value == "2026-12-31"
    assert sheet["A3"].value == "Certificates"
    assert sheet["B3"].value == "Certificate ABC"
    assert sheet["C3"].value == "End of Extended Vendor Support"
    assert sheet["D3"].value == "2026-10-01"


def test_build_expiration_workbook_applies_status_colors_for_expired_and_active_dates():
    today = date.today()
    expired_date = today.replace(year=today.year - 1).isoformat()
    active_date = today.replace(year=today.year + 1).isoformat()
    records = [
        ExpirationRecord(
            source_section="Expired",
            name="Old component",
            expiration_date=expired_date,
            milestone="End of Standard Vendor Support",
        ),
        ExpirationRecord(
            source_section="Active",
            name="Future component",
            expiration_date=active_date,
            milestone="",
        ),
    ]

    payload = build_expiration_workbook(records)

    workbook = load_workbook(filename=BytesIO(payload))
    sheet = workbook.active

    expired_fill = sheet["A2"].fill.fgColor.rgb
    active_fill = sheet["A3"].fill.fgColor.rgb

    assert expired_fill == "00F6E7A1"
    assert sheet["B2"].fill.fgColor.rgb == "00F6E7A1"
    assert sheet["C2"].fill.fgColor.rgb == "00F6E7A1"
    assert sheet["D2"].fill.fgColor.rgb == "00F6E7A1"
    assert active_fill == "00CFE8C6"
    assert sheet["B3"].fill.fgColor.rgb == "00CFE8C6"
    assert sheet["C3"].fill.fgColor.rgb == "00CFE8C6"
    assert sheet["D3"].fill.fgColor.rgb == "00CFE8C6"


def test_build_consolidated_workbook_generates_base_client_view_and_uncataloged_sheets():
    payload = build_consolidated_workbook(
        ConsolidatedWorkbookData(
            clients=[("Cliente A", "2026-04"), ("Cliente B", "2026-04")],
            records=[
                ConsolidatedExpiration(
                    client="Cliente A",
                    period="2026-04",
                    component="SAP Kernel",
                    detected_name="SAP Kernel Release",
                    milestone="End of Standard Vendor Support",
                    expiration_date="2026-12-31",
                    source_section="SAP Kernel Release",
                    source_filename="cliente-a.pdf",
                    is_cataloged=True,
                ),
                ConsolidatedExpiration(
                    client="Cliente A",
                    period="2026-04",
                    component="SAP Solution Manager",
                    detected_name="SAP Solution Manager 7.2",
                    milestone="",
                    expiration_date="2027-12-31",
                    source_section="Maintenance",
                    source_filename="cliente-a.pdf",
                    is_cataloged=True,
                ),
                ConsolidatedExpiration(
                    client="Cliente A",
                    period="2026-04",
                    component="SAP Cloud Connector",
                    detected_name="SAP Cloud Connector",
                    milestone="",
                    expiration_date="2027-01-31",
                    source_section="Custom",
                    source_filename="cliente-a.pdf",
                    is_cataloged=False,
                ),
            ],
        )
    )

    workbook = load_workbook(filename=BytesIO(payload))

    assert workbook.sheetnames == ["Base", "VistaClientes", "ComponentesNoCatalogados"]

    base = workbook["Base"]
    assert [base.cell(row=1, column=column).value for column in range(1, 4)] == [
        "Cliente",
        "Componente",
        "FechaVencimiento",
    ]
    assert [base.cell(row=2, column=column).value for column in range(1, 4)] == [
        "Cliente A",
        "SAP Kernel",
        "2026-12-31",
    ]

    client_view = workbook["VistaClientes"]
    assert [client_view.cell(row=1, column=column).value for column in range(1, 13)] == [
        "Cliente",
        "Periodo",
        "SAP Product Version",
        "SAP NetWeaver",
        "SAP Solution Manager",
        "SAP Fiori",
        "SAP Kernel",
        "Database",
        "Operating System",
        "Support Package Stack",
        "Certificates",
        "Otros componentes",
    ]
    assert client_view["A2"].value == "Cliente A"
    assert client_view["B2"].value == "2026-04"
    assert client_view["E2"].value == "2027-12-31"
    assert client_view["G2"].value == "2026-12-31 (End of Standard Vendor Support)"
    assert client_view["L2"].value == "SAP Cloud Connector: 2027-01-31"
    assert client_view["A3"].value == "Cliente B"
    assert client_view["G3"].value is None

    uncataloged = workbook["ComponentesNoCatalogados"]
    assert [uncataloged.cell(row=1, column=column).value for column in range(1, 6)] == [
        "NombreDetectado",
        "Cliente",
        "Periodo",
        "FechaVencimiento",
        "FuenteEWA",
    ]
    assert [uncataloged.cell(row=2, column=column).value for column in range(1, 6)] == [
        "SAP Cloud Connector",
        "Cliente A",
        "2026-04",
        "2027-01-31",
        "cliente-a.pdf",
    ]
