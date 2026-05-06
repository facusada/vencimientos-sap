from io import BytesIO
import json

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.models.expiration import AiUsageMetrics
from app.services.document_intelligence import DocumentIntelligenceProvider
from app.services.ai_usage_repository import AiUsageRepository
from app.services.ai_usage_repository import get_ai_usage_repository
from app.services.ewa_analysis_service import get_document_intelligence_provider


client = TestClient(app)


class StubAiUsageRepository(AiUsageRepository):
    def __init__(self) -> None:
        self.saved_usages = []

    def save_usages(self, usages) -> None:
        self.saved_usages.extend(usages)


class StubDocumentIntelligenceProvider(DocumentIntelligenceProvider):
    def extract_expirations(self, text: str) -> list[dict[str, str]]:
        assert "supported until 02.2027" in text
        return [
            {"nombre": "SAP Product Version", "fecha": "02.2027"},
            {"nombre": "Kernel", "fecha": "2026-12-31", "hito": "End of Standard Vendor Support"},
            {"nombre": "Kernel", "fecha": "2026-12-31", "hito": "End of Standard Vendor Support"},
        ]


def test_post_ewa_analyze_returns_excel_file_for_pdf(monkeypatch):
    app.dependency_overrides[get_document_intelligence_provider] = (
        lambda: StubDocumentIntelligenceProvider()
    )
    monkeypatch.setattr(
        "app.services.ewa_analysis_service.extract_text",
        lambda filename, payload: (
            "SAP Product Version is supported until 02.2027.\nKernel expires on 2026-12-31."
            if filename == "ewa.pdf" and payload == b"fake-pdf"
            else ""
        ),
    )

    try:
        response = client.post(
            "/ewa/analyze",
            files={
                "file": (
                    "ewa.pdf",
                    b"fake-pdf",
                    "application/pdf",
                )
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=ewa-expirations.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(filename=BytesIO(response.content))
    sheet = workbook.active

    assert sheet["A1"].value == "Seccion"
    assert sheet["B1"].value == "Nombre"
    assert sheet["C1"].value == "Hito"
    assert sheet["D1"].value == "Fecha"
    assert sheet["A2"].value is None
    assert sheet["B2"].value == "SAP Product Version"
    assert sheet["C2"].value is None
    assert sheet["D2"].value == "2027-02-28"
    assert sheet["A3"].value is None
    assert sheet["B3"].value == "Kernel"
    assert sheet["C3"].value == "End of Standard Vendor Support"
    assert sheet["D3"].value == "2026-12-31"


def test_post_api_ewa_analyze_returns_excel_file_for_pdf(monkeypatch):
    app.dependency_overrides[get_document_intelligence_provider] = (
        lambda: StubDocumentIntelligenceProvider()
    )
    monkeypatch.setattr(
        "app.services.ewa_analysis_service.extract_text",
        lambda filename, payload: (
            "SAP Product Version is supported until 02.2027.\nKernel expires on 2026-12-31."
            if filename == "ewa.pdf" and payload == b"fake-pdf"
            else ""
        ),
    )

    try:
        response = client.post(
            "/api/ewa/analyze",
            files={
                "file": (
                    "ewa.pdf",
                    b"fake-pdf",
                    "application/pdf",
                )
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=ewa-expirations.xlsx" in response.headers["content-disposition"]


def test_post_ewa_analyze_rejects_unsupported_extension():
    response = client.post(
        "/ewa/analyze",
        files={"file": ("ewa.csv", b"irrelevant", "text/csv")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Unsupported file type"


def test_post_ewa_analyze_returns_clear_error_when_ai_finds_nothing(monkeypatch):
    class EmptyProvider(DocumentIntelligenceProvider):
        def extract_expirations(self, text: str) -> list[dict[str, str]]:
            return []

    app.dependency_overrides[get_document_intelligence_provider] = lambda: EmptyProvider()
    monkeypatch.setattr(
        "app.services.ewa_analysis_service.extract_text",
        lambda filename, payload: (
            "This EWA contains recommendations but no support dates."
            if filename == "ewa.pdf" and payload == b"fake-pdf"
            else ""
        ),
    )

    try:
        response = client.post(
            "/ewa/analyze",
            files={
                "file": (
                    "ewa.pdf",
                    b"fake-pdf",
                    "application/pdf",
                )
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 400
    assert response.json()["detail"] == "No se detectaron fechas de vencimiento en el EWA enviado."


def test_post_ewa_analyze_rejects_non_pdf_input():
    response = client.post(
        "/ewa/analyze",
        files={
            "file": (
                "ewa.docx",
                b"legacy-docx",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Unsupported file type"


def test_post_ewa_consolidate_returns_monthly_workbook_for_multiple_pdfs(monkeypatch):
    class ConsolidationProvider(DocumentIntelligenceProvider):
        def extract_expirations(self, text: str) -> list[dict[str, str]]:
            if "cliente-a" in text:
                return [
                    {
                        "nombre": "SAP Kernel Release",
                        "fecha": "2026-12-31",
                        "hito": "End of Standard Vendor Support",
                    },
                    {"nombre": "SAP Cloud Connector", "fecha": "2027-01-31"},
                ]
            return [{"nombre": "SAP Fiori Front-End Server", "fecha": "31.12.2027"}]

        def extract_expirations_with_usage(
            self,
            text: str,
        ) -> tuple[list[dict[str, str]], AiUsageMetrics | None]:
            items = self.extract_expirations(text)
            if "cliente-a" in text:
                return items, AiUsageMetrics(input_tokens=1100, output_tokens=140, total_tokens=1240)
            return items, AiUsageMetrics(input_tokens=950, output_tokens=90, total_tokens=1040)

    usage_repository = StubAiUsageRepository()
    app.dependency_overrides[get_document_intelligence_provider] = lambda: ConsolidationProvider()
    app.dependency_overrides[get_ai_usage_repository] = lambda: usage_repository
    monkeypatch.setattr(
        "app.services.ewa_analysis_service.extract_text",
        lambda filename, payload: (
            f"text for cliente-a {filename}" if filename == "a.pdf" else f"text for cliente-b {filename}"
        ),
    )

    try:
        response = client.post(
            "/ewa/consolidate",
            data={
                "period": "2026-04",
                "clients": ["Cliente A", "Cliente B"],
            },
            files=[
                ("files", ("a.pdf", b"a-pdf", "application/pdf")),
                ("files", ("b.pdf", b"b-pdf", "application/pdf")),
            ],
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=ewa-consolidated.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(filename=BytesIO(response.content))
    assert workbook.sheetnames == ["Base"]
    assert workbook["Base"]["A2"].value == "Cliente A"
    assert workbook["Base"]["B2"].value == "SAP Kernel Release"
    assert workbook["Base"]["C2"].value == "2026-12-31"
    assert "x-ewa-no-results" not in response.headers
    assert [
        (item.client, item.input_tokens, item.output_tokens, item.total_tokens)
        for item in usage_repository.saved_usages
    ] == [
        ("Cliente A", 1100, 140, 1240),
        ("Cliente B", 950, 90, 1040),
    ]


def test_post_ewa_consolidate_reports_ewa_without_expiration_results(monkeypatch):
    class PartialConsolidationProvider(DocumentIntelligenceProvider):
        def extract_expirations(self, text: str) -> list[dict[str, str]]:
            if "cliente-a" in text:
                return [{"nombre": "Kernel", "fecha": "2026-12-31"}]
            return []

    usage_repository = StubAiUsageRepository()
    app.dependency_overrides[get_document_intelligence_provider] = lambda: PartialConsolidationProvider()
    app.dependency_overrides[get_ai_usage_repository] = lambda: usage_repository
    monkeypatch.setattr(
        "app.services.ewa_analysis_service.extract_text",
        lambda filename, payload: (
            f"text for cliente-a {filename}" if filename == "a.pdf" else f"text for cliente-b {filename}"
        ),
    )

    try:
        response = client.post(
            "/ewa/consolidate",
            data={
                "period": "2026-04",
                "clients": ["Cliente A", "Cliente B"],
            },
            files=[
                ("files", ("a.pdf", b"a-pdf", "application/pdf")),
                ("files", ("b.pdf", b"b-pdf", "application/pdf")),
            ],
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200

    workbook = load_workbook(filename=BytesIO(response.content))
    assert workbook.sheetnames == ["Base"]
    assert json.loads(response.headers["x-ewa-no-results"]) == [
        {
            "client": "Cliente B",
            "period": "2026-04",
            "filename": "b.pdf",
            "reason": "Sin vencimientos detectados",
        }
    ]
    assert [
        (item.client, item.input_tokens, item.output_tokens, item.total_tokens)
        for item in usage_repository.saved_usages
    ] == [
        ("Cliente A", None, None, None),
        ("Cliente B", None, None, None),
    ]


def test_post_api_ewa_consolidate_alias_returns_workbook(monkeypatch):
    class ConsolidationProvider(DocumentIntelligenceProvider):
        def extract_expirations(self, text: str) -> list[dict[str, str]]:
            return [{"nombre": "Kernel", "fecha": "2026-12-31"}]

    usage_repository = StubAiUsageRepository()
    app.dependency_overrides[get_document_intelligence_provider] = lambda: ConsolidationProvider()
    app.dependency_overrides[get_ai_usage_repository] = lambda: usage_repository
    monkeypatch.setattr(
        "app.services.ewa_analysis_service.extract_text",
        lambda filename, payload: "Kernel expires on 2026-12-31.",
    )

    try:
        response = client.post(
            "/api/ewa/consolidate",
            data={"period": "2026-04", "clients": ["Cliente A"]},
            files=[("files", ("a.pdf", b"a-pdf", "application/pdf"))],
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert "attachment; filename=ewa-consolidated.xlsx" in response.headers["content-disposition"]


def test_post_ewa_consolidate_rejects_client_file_count_mismatch():
    response = client.post(
        "/ewa/consolidate",
        data={"period": "2026-04", "clients": ["Cliente A"]},
        files=[
            ("files", ("a.pdf", b"a-pdf", "application/pdf")),
            ("files", ("b.pdf", b"b-pdf", "application/pdf")),
        ],
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "La cantidad de clientes debe coincidir con la cantidad de archivos."


def test_post_ewa_consolidate_rejects_invalid_period():
    response = client.post(
        "/ewa/consolidate",
        data={"period": "202604", "clients": ["Cliente A"]},
        files=[("files", ("a.pdf", b"a-pdf", "application/pdf"))],
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "El periodo debe tener formato YYYY-MM."
