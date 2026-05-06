from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.expiration import ConsolidatedWorkbookData
from app.models.expiration import ExpirationRecord

EXPIRED_FILL = PatternFill(fill_type="solid", fgColor="F6E7A1")
ACTIVE_FILL = PatternFill(fill_type="solid", fgColor="CFE8C6")
BODY_FONT = Font(color="1C1B18")
HEADER_FONT = Font(bold=True, color="1C1B18")


def build_expiration_workbook(records: list[ExpirationRecord]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expirations"
    sheet.append(["Seccion", "Nombre", "Hito", "Fecha"])

    for record in records:
        sheet.append([record.source_section, record.name, record.milestone, record.expiration_date])
        _apply_status_style(sheet, sheet.max_row, record.expiration_date)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_consolidated_workbook(data: ConsolidatedWorkbookData) -> bytes:
    workbook = Workbook()
    base_sheet = workbook.active
    base_sheet.title = "Base"
    _build_base_sheet(base_sheet, data)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_base_sheet(sheet, data: ConsolidatedWorkbookData) -> None:
    sheet.append(["Cliente", "Componente", "FechaVencimiento"])
    _style_header(sheet, 3)

    for record in data.records:
        sheet.append(
            [
                record.client,
                record.component,
                record.expiration_date,
            ]
        )
        _apply_status_style(sheet, sheet.max_row, record.expiration_date, columns=("A", "B", "C"))


def _style_header(sheet, columns_count: int) -> None:
    for column_index in range(1, columns_count + 1):
        sheet.cell(row=1, column=column_index).font = HEADER_FONT


def _apply_status_style(
    sheet,
    row_index: int,
    expiration_date: str,
    columns: tuple[str, ...] = ("A", "B", "C", "D"),
) -> None:
    target_fill = _resolve_fill(expiration_date)

    for column in columns:
        cell = sheet[f"{column}{row_index}"]
        cell.fill = target_fill
        cell.font = BODY_FONT


def _resolve_fill(expiration_date: str) -> PatternFill:
    try:
        normalized_date = date.fromisoformat(expiration_date)
    except ValueError:
        return ACTIVE_FILL

    return EXPIRED_FILL if normalized_date < date.today() else ACTIVE_FILL
