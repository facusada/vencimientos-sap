from io import BytesIO

from openpyxl import Workbook
from openpyxl import load_workbook

BASE_SHEET_NAME = "Base"
BASE_HEADER = ("Cliente", "Componente", "FechaVencimiento")


def merge_ewa_workbooks(workbooks: list[tuple[str, bytes]]) -> bytes:
    if len(workbooks) < 2:
        raise ValueError("El merge requiere al menos 2 workbooks.")

    merged_rows: list[tuple[str, str, str]] = []

    for filename, payload in workbooks:
        merged_rows.extend(_extract_base_rows(filename, payload))

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = BASE_SHEET_NAME
    output_sheet.append(list(BASE_HEADER))

    for row in merged_rows:
        output_sheet.append(list(row))

    buffer = BytesIO()
    output_workbook.save(buffer)
    return buffer.getvalue()


def _extract_base_rows(filename: str, payload: bytes) -> list[tuple[str, str, str]]:
    workbook = load_workbook(filename=BytesIO(payload), data_only=True)

    if BASE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"El workbook {filename} debe incluir la hoja Base.")

    sheet = workbook[BASE_SHEET_NAME]
    header = tuple(sheet.cell(row=1, column=column).value for column in range(1, 4))
    if header != BASE_HEADER:
        raise ValueError(
            f"El workbook {filename} debe incluir las columnas requeridas: "
            f"{', '.join(BASE_HEADER)}."
        )

    rows: list[tuple[str, str, str]] = []
    for values in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        normalized = tuple("" if value is None else str(value).strip() for value in values)
        if not any(normalized):
            continue
        rows.append(normalized)

    return rows
